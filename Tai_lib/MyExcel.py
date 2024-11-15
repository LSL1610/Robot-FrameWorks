from time import sleep
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import listdir, path
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side
from robot.libraries.OperatingSystem import OperatingSystem
from robot.api.deco import keyword
from openpyxl.utils.cell import column_index_from_string

myOS = OperatingSystem()

curMonth = datetime.now().strftime("%B_%Y")  # November_2021
curDate = datetime.now().strftime("%d_%b")  # 26_Nov
curTime = datetime.now().strftime("%H")  # 01 - 24
#curMinute = datetime.now().strftime("%M")  # 00 - 59
#curAP = datetime.now().strftime("%p")  # AM/PM

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# --------------------------


class MyExcel:
    def __init__(self) -> None:
        pass

    # ---------------------- create an "Excel raw file" and copy data from "Excel result form"
    # with option "folder tool name"
    @keyword('Create File Excel Raw With Format')
    def Optional_creatExcelFileRawWithFormat(self, folder_tool_path, result_format_path, report_name = "raw_results.xlsx"):
        """Arguments: Folder tool path | Result format path

        Creates the specified "raw_results.xlsx" excel file in the RESULTS directory and copy data from file "result_format".

        Example:
        | Create File Excel Raw | Folder_Path | Folder_Path\\Format_file_name.xlsx  |   report_name.xlsx
        """
        # Directory path of "Excel raw" and "Excel format"
        myOS.create_directory(folder_tool_path)
        ex_result_raw = path.join(folder_tool_path, report_name)
        ex_result_format = result_format_path

        # Create an Excel file with name "raw_results.xlsx"
        wb_raw = Workbook()
        wb_raw.save(ex_result_raw)
        wb_raw = load_workbook(ex_result_raw)
        ex_result_raw_sheet = wb_raw.active

        # Copy all data report from "Excel format" to "Excel raw"
        wb_format = load_workbook(ex_result_format)
        wb_format_sheet = wb_format.active

        for i in range(1, wb_format_sheet.max_row + 1):
            for j in range(1, wb_format_sheet.max_column + 1):
                ex_result_raw_sheet.cell(row=i, column=j).value = wb_format_sheet.cell(row=i, column=j).value
        wb_raw.save(ex_result_raw)
        print('Created file raw_result.xlsx')

    @keyword('Create File Excel Raw Without Format')
    def Optional_creatExcelFileRaw(self, folder_tool_path, report_name = "raw_results.xlsx"):
        """Arguments: Folder tool path | Result format path

        Creates the specified "raw_results.xlsx" excel file in the RESULTS directory and copy data from file "result_format".

        Example:
        | Create File Excel Raw | Folder_Path |   report_name.xlsx
        """
        # Directory path of "Excel raw" and "Excel format"
        myOS.create_directory(folder_tool_path)
        result_raw = path.join(folder_tool_path, report_name)

        # Create an Excel file with name "raw_results.xlsx"
        wb_raw = Workbook()
        wb_raw.save(result_raw)
        print('Created file raw')
    # ---------------------- create monthly folder and daily report with options "folder_report_path" and "report_name"
    @keyword('Create Monthly Directory And Daily Report')
    def Optional_createFolderMonthlyReport(self, folder_report_path, report_name):
        """Arguments: Path | Report file name

        Creates the specified current month directory and daily report file of the project.

        *Note: The argument report_name must be written as the unique name of the tool.The keyword will generate the current date value and concatenate it with the given name, and save it in .xlsx format.

        Example:
        | Create Monthly Directory And Daily Report | Folder_Path   | Report_Tool_A  |
        
        Expected output value:
        | G:\\My Drive\\PROJECT_Owner\\Tool_A\\December_2021\\Report_A_01_Dec.xlsx |
        """
        # Chon 1 máy chạy tạo và quét folder của tháng mỗi ngày (hiện tại là VNPT)
        # Đường dẫn đến thư mục Gốc cần chứa thư mục Tháng của tool
        folderReportPath = folder_report_path
        # Tên của file report muốn đặt kết hợp với ngày hiện tại với định dạng mong muốn ( hiện tại là .xlsx )
        reportName = str(report_name) + "_" + curDate + ".xlsx"
        # Đường dẫn đầy đủ của file report, dùng cho việc lưu file report phía dưới
        ExcelFilePath = path.join(folder_report_path, curMonth, reportName)
        pathFolderMonth = path.join(folder_report_path, curMonth)
        try:
            # Kiểm tra tồn tại của thư mục Tháng hiện hành trong thư mục Gốc
            if curMonth not in listdir(folderReportPath):
                print('Current month directory NOT FOUND.')
                myOS.create_directory(pathFolderMonth)
                print('Created folder ' + curMonth)
            # Kiểm tra tồn tại file daily report trong thư mục Tháng
            if reportName not in listdir(pathFolderMonth):
                wb = Workbook()
                wb.save(ExcelFilePath)
                print('Created daily report file ' + reportName)
            # Kiểm tra tồn tại của sheet mốc giờ report của daily report
            wb = load_workbook(ExcelFilePath)
            ws = wb.sheetnames
            if curTime not in str(ws):
                wb.create_sheet(str(curTime), 0)
                wb.save(ExcelFilePath)
                print('Created sheet name: ' + str(curTime))
        except Exception as e:
            print(e)

    @keyword('Create Monthly Directory And Daily Report 2')
    def Optional_createFolderMonthlyReport2(self, folder_report_path, report_name):
        """Arguments: Path | Report file name

        Creates the specified current month directory and daily report file of the project.

        *Note: The argument report_name must be written as the unique name of the tool.The keyword will save it in .xlsx format.

        Example:
        | Create Monthly Directory And Daily Report | Folder_Path   | Report_Tool_A  |
        
        Expected output value:
        | G:\\My Drive\\PROJECT_Owner\\Tool_A\\December_2021\\Report_A.xlsx |
        """
        # Chon 1 máy chạy tạo và quét folder của tháng mỗi ngày (hiện tại là VNPT)
        # Đường dẫn đến thư mục Gốc cần chứa thư mục Tháng của tool
        folderReportPath = folder_report_path
        reportName = str(report_name) + ".xlsx"
        # Tên của file report muốn đặt kết hợp với ngày hiện tại với định dạng mong muốn ( hiện tại là .xlsx )
        # Đường dẫn đầy đủ của file report, dùng cho việc lưu file report phía dưới
        ExcelFilePath = path.join(folder_report_path, curMonth, reportName)
        pathFolderMonth = path.join(folder_report_path, curMonth)
        try:
            # Kiểm tra tồn tại của thư mục Tháng hiện hành trong thư mục Gốc
            if curMonth not in listdir(folderReportPath):
                print('Current month directory NOT FOUND.')
                myOS.create_directory(pathFolderMonth)
                print('Created folder ' + curMonth)
            # Kiểm tra tồn tại file daily report trong thư mục Tháng
            if reportName not in listdir(pathFolderMonth):
                wb = Workbook()
                wb.save(ExcelFilePath)
                print('Created daily report file ' + reportName)
            # Kiểm tra tồn tại của sheet mốc giờ report của daily report
            wb = load_workbook(ExcelFilePath)
            ws = wb.sheetnames
            if curDate not in str(ws):
                wb.create_sheet(str(curDate), 0)
                wb.save(ExcelFilePath)
                print('Created sheet name: ' + str(curDate))
        except Exception as e:
            print(e)


    # ---------------------- Copy all results from raw report file to final report
    @keyword('Copy To Final Report')
    def Optional_copyFinalReportFile(self, local_report_path, name_report_raw, final_report_path, report_name, border = "thin"):
        """Arguments: Local file path | Final file path | Report_Name   | Border

        Copy all data on local report file raw to final file report of the Project.
        Border value must be one of 'dashDotDot', 'medium', 'dotted', 'slantDashDot', 'thin', 'hair', 
        'mediumDashDotDot', 'dashDot', 'double', 'mediumDashed', 'dashed', 'mediumDashDot', 'thick'

        *Note: 
         The argument report_name must be written as the unique name of the tool. 
         The keyword will generate the current date value and concatenate it with the given name.
        
        Example:
        | Copy To Final Report  | D:\\PROJECT_Owner\\Tool_A | G:\\My Drive\\PROJECT_Owner\\Tool_A   | Report_Name  |
        """
        final_border = Border(left=Side(style= border),
                     right=Side(style= border),
                     top=Side(style= border),
                     bottom=Side(style= border))

        sleep(30)
        # Load raw_results.xlsx
        result_raw_path = path.join(local_report_path, name_report_raw)
        wb_raw = load_workbook(result_raw_path)
        wb_raw_sheet = wb_raw.active
        mr = wb_raw_sheet.max_row
        mc = wb_raw_sheet.max_column

        # Load daily_report
        reportName = str(report_name) + "_" + curDate + ".xlsx"
        FinalExcelFilePath = path.join(final_report_path, curMonth, reportName)
        wb_final = load_workbook(FinalExcelFilePath) # remove result_path duplicated with FinalExcelFilePath
        wb_final_sheet = wb_final.active
        mr1 = wb_final_sheet.max_row
        mc1 = wb_final_sheet.max_column
        
        # Copy results of file "report_raw" to file "report_final"
        last_col = mc1
        print("last col:", last_col)
        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                wb_raw_sheet_value = wb_raw_sheet.cell(row=r, column=c).value
                # last_col = 1: Report's data is None
                if last_col == 1:
                    wb_final_sheet.cell(row=r, column=c).value = wb_raw_sheet_value
                    wb_final_sheet.cell(row=r, column=c).border = final_border
                    wb_final_sheet.cell(row=r, column=c).comment = wb_raw_sheet.cell(row=r, column=c).comment
                # last_col > 1: Report has data, paste new results to the next last column
                else:
                    wb_final_sheet.cell(row=r, column=c + last_col).value = wb_raw_sheet_value
                    wb_final_sheet.cell(row=r, column=c + last_col).border = final_border
                    wb_final_sheet.cell(row=r, column=c + last_col).comment = wb_raw_sheet.cell(row=r, column=c).comment
        wb_final.save(FinalExcelFilePath)
        print('Copy data to final report completed')

    @keyword('Copy To Final Report Without Current Date')
    def Optional_copyFinalReportFileWithoutCurrentDate(self, local_report_path, final_report_path, report_name, border = "thin"):
        """Arguments: Local file path | Final file path | Report_Name   | Border

        Copy all data on local report file raw to final file report of the Project.
        Border value must be one of 'dashDotDot', 'medium', 'dotted', 'slantDashDot', 'thin', 'hair', 
        'mediumDashDotDot', 'dashDot', 'double', 'mediumDashed', 'dashed', 'mediumDashDot', 'thick'

        *Note: 
         The argument report_name must be written as the unique name of the tool. 
        
        Example:
        | Copy To Final Report  | D:\\PROJECT_Owner\\Tool_A | G:\\My Drive\\PROJECT_Owner\\Tool_A   | Report_Name  |
        """
        final_border = Border(left=Side(style= border),
                     right=Side(style= border),
                     top=Side(style= border),
                     bottom=Side(style= border))

        sleep(30)
        # Load raw_results.xlsx
        result_raw_path = path.join(local_report_path, "raw_results.xlsx")
        wb_raw = load_workbook(result_raw_path)
        wb_raw_sheet = wb_raw.active
        mr = wb_raw_sheet.max_row
        mc = wb_raw_sheet.max_column

        # Load daily_report
        reportName = str(report_name) + ".xlsx"
        FinalExcelFilePath = path.join(final_report_path, curMonth, reportName)
        wb_final = load_workbook(FinalExcelFilePath) # remove result_path duplicated with FinalExcelFilePath
        wb_final_sheet = wb_final.active
        mr1 = wb_final_sheet.max_row
        mc1 = wb_final_sheet.max_column
        
        # Copy results of file "report_raw" to file "report_final"
        last_col = mc1
        print("last col:", last_col)
        for r in range(1, mr + 1):
            for c in range(1, mc + 1):
                wb_raw_sheet_value = wb_raw_sheet.cell(row=r, column=c).value
                # last_col = 1: Report's data is None
                if last_col == 1:
                    wb_final_sheet.cell(row=r, column=c).value = wb_raw_sheet_value
                    wb_final_sheet.cell(row=r, column=c).border = final_border
                    wb_final_sheet.cell(row=r, column=c).comment = wb_raw_sheet.cell(row=r, column=c).comment
                # last_col > 1: Report has data, paste new results to the next last column
                else:
                    wb_final_sheet.cell(row=r, column=c + last_col).value = wb_raw_sheet_value
                    wb_final_sheet.cell(row=r, column=c + last_col).border = final_border
                    wb_final_sheet.cell(row=r, column=c + last_col).comment = wb_raw_sheet.cell(row=r, column=c).comment
        wb_final.save(FinalExcelFilePath)
        print('Copy data to final report completed')
    # ---------------------- Fill Color with conditions
    @keyword('Fill Value Color')
    def FillValueColor(self, final_report_path, report_name, condition, fgColor):
        """Arguments: Final report path | Report name | Condition | fgColor

        Fill the specified color with given condition.
        Note: Condition's color must be aRGB hex values.

        Example:
        | ${condition}      | Create List                | PASSED        | Time          | ABC       |
        | Fill Value Color  | G:\\My Drive\\Tool_Report  | Report_Name   | ${condition}  | 00FFFFFF  |
        """
        # Load daily_report
        reportName = str(report_name) + "_" + curDate + ".xlsx"
        FinalExcelFilePath = path.join(final_report_path, curMonth, reportName)
        wb_final = load_workbook(FinalExcelFilePath) # remove result_path duplicated with FinalExcelFilePath
        wb_final_sheet = wb_final.active

        for i in range(1, wb_final_sheet.max_row + 1):
            for j in range(1, wb_final_sheet.max_column + 1):
                value = wb_final_sheet.cell(row=i, column=j).value
                if str(value) is not None and str(condition) in str(value):
                    wb_final_sheet.cell(row=i, column=j).fill = PatternFill("solid", fgColor = fgColor)
                elif str(value) is not None and str(value) in condition:
                    wb_final_sheet.cell(row=i, column=j).fill = PatternFill("solid", fgColor = fgColor)
                if value is None:
                    wb_final_sheet.cell(row=i, column=j).fill = PatternFill("solid", fgColor="FFFF0000")

        wb_final.save(FinalExcelFilePath)

    @keyword('Resize Column')
    def ResizeColumn(self,final_report_path, report_name, condition, size):
        """Arguments: Final report path | Report name | Condition | Size

        Resize specified column with given condition.
        Note: Size's value must be float number values.

        Example:
        | ${condition}      | Create List                | PASSED        | Time          | Domain   |
        | Fill Value Color  | G:\\My Drive\\Tool_Report  | Report_Name   | ${condition}  | 15       |
        """
        # Load daily_report
        reportName = str(report_name) + "_" + curDate + ".xlsx"
        FinalExcelFilePath = path.join(final_report_path, curMonth, reportName)
        #result_final_path = FinalExcelFilePath
        wb_final = load_workbook(FinalExcelFilePath)
        wb_final_sheet = wb_final.active
        mr1 = wb_final_sheet.max_row + 1
        mc1 = wb_final_sheet.max_column + 1

        for c in wb_final_sheet.iter_cols(min_col=1, max_col= mc1):
            for cell in c:
                for value in condition:
                    if value in str(cell.value):
                        col_letter = cell.column_letter
                        wb_final_sheet.column_dimensions[col_letter].width = size

        wb_final.save(FinalExcelFilePath)

    @keyword('Freeze Panels Excel')
    def freeze_panels(self,final_report_path, report_name: str, col_name: str='B2'):
        reportName = str(report_name) + "_" + curDate + ".xlsx"
        FinalExcelFilePath = path.join(final_report_path, curMonth, reportName)
        wb = load_workbook(FinalExcelFilePath)
        ws = wb.active
        # want to freeze panel, default both row 1 and column A
        c = ws[f'{col_name}']
        ws.freeze_panes = c
        wb.save(FinalExcelFilePath)

    @keyword('Transfer Failed Message To Commnent')
    def TransferMessageToComment(self, final_report_path: str, report_name: str, network: str) -> None:
        """Arguments: final_report_path | report_name | network

        Transfer the following keyword failed message from "FAILED" to a comment.
        Network Name must be one of: VIETTEL, VNPT, MOBI, FPT.

        Example:
        |   Transfer Failed Message To Commnent   |   final_report_path  | report_name   |  VNPT    |
        """
        daily_name = report_name + "_" + curDate + ".xlsx"
        final_path = path.join(final_report_path, curMonth, daily_name)
        wb_final = load_workbook(final_path)
        ws_final = wb_final.active
        mr = ws_final.max_row
        mc = ws_final.max_column

        for c in ws_final.iter_cols(min_col=1, max_col=mc):
            for cell in c:
                if network in str(cell.value):
                    col_letter = cell.column_letter
                    icolumn = column_index_from_string(col_letter)
                    for i in range(1, mr + 1):
                        cell_value = ws_final.cell(row=i, column=icolumn).value
                        if ws_final.cell(row=i, column=icolumn).value is None:
                            pass
                        elif "FAILED - " in cell_value:
                            cell_value = ws_final.cell(row=i, column=icolumn).value
                            txt_fail = cell_value[0:6]
                            trim_str = cell_value.strip(txt_fail)
                            comment = Comment(trim_str, "")
                            ws_final[str(col_letter) + str(i)].comment = comment
                            ws_final.cell(row=i, column=icolumn).value = txt_fail
        wb_final.save(final_path)
 
# --------- debug

#me = MyExcel()
#me.TransferMessageToComment('D:\\PROJECT_DAT\\Check_Alive_B1_12_Brands\\RESULTS\\raw_results.xlsx','VIETTEL')
#me.freeze_panels(r'G:\My Drive\PROJECT_DAT\1_my_libs\my_debug\Iframe_Results.xlsx')
#me.Optional_creatExcelFileRaw(r'D:\PROJECT_DAT\Check_Alive_B1_12_Brands', r'D:\PROJECT_DAT\Check_Alive_B1_12_Brands\RESULTS\result_format.xlsx')
#me.Optional_createFolderMonthlyReport(r'D:\PROJECT_DAT\task_iframe', r'Iframe_Results_Demo')
#me.TransferMessageToComment(r'D:\PROJECT_DAT\VT\Check_iframe\RESULTS\raw_results.xlsx', 'VIETTEL')
#me.Optional_copyFinalReportFile(r'D:\PROJECT_DAT\Check_Alive_B1_12_Brands', r'D:\PROJECT_DAT\task_iframe', r'Iframe_Results_Demo')
#list1 = ["Domain", "VIETTEL"]
#me.FillValueColor(r'D:\PROJECT_DAT\task_iframe', r'Iframe_Results_Demo', list1, 'A3E4D7')
#me.ResizeColumn(r'D:\PROJECT_DAT\task_iframe', r'Iframe_Results_Demo', list1, 15)
